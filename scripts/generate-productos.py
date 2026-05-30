"""Genera public/productos.json desde Codigos Productos y datos.xlsx"""
import json
import pathlib
import re
import shutil

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX_SRC = ROOT / 'Codigos Productos y datos.xlsx'
JSON_OUT = ROOT / 'public' / 'productos.json'
XLSX_PUBLIC = ROOT / 'public' / 'codigos-productos-y-datos.xlsx'


def limpiar_codigo(codigo):
    """Convierte artefactos Excel (_x0002_ etc.) en caracteres legibles."""
    s = str(codigo).strip()

    def reemplazar_escape(match):
        try:
            char = chr(int(match.group(1), 16))
            # Caracteres de control → guión bajo (habitual en códigos de producto)
            return '_' if ord(char) < 32 else char
        except ValueError:
            return '_'

    s = re.sub(r'_x([0-9a-fA-F]{4})_', reemplazar_escape, s)
    return s


def limpiar_nombre(nombre):
    """Unifica saltos de línea y espacios del Excel."""
    return ' '.join(str(nombre or '').split())


def main():
    if not XLSX_SRC.exists():
        raise FileNotFoundError(f'No se encontró {XLSX_SRC}')

    wb = openpyxl.load_workbook(XLSX_SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    productos = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
        codigo = row[1]
        if not codigo:
            continue

        nombre = row[6]
        unidades = row[7]
        volumen = row[8]
        fob = row[9]
        derechos = row[10]

        if unidades is None or volumen is None or fob is None or derechos is None:
            continue

        productos.append({
            'id': f'row-{row_idx}',
            'codigo': limpiar_codigo(codigo),
            'nombre': limpiar_nombre(nombre),
            'unidadesPorCaja': float(unidades),
            'volumenPorCaja': float(volumen),
            'precioFOBUnitario': float(fob),
            'pctDerechos': round(float(derechos) * 100, 4),
        })

    wb.close()

    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    with JSON_OUT.open('w', encoding='utf-8') as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)

    shutil.copy2(XLSX_SRC, XLSX_PUBLIC)

    print(f'Generados {len(productos)} productos -> {JSON_OUT}')
    print(f'Excel copiado -> {XLSX_PUBLIC}')


if __name__ == '__main__':
    main()
